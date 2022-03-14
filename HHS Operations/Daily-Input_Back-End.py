# !Python3                                              A3R0NA$
import openpyxl as op

# Open Project Tracker and Daily Input, load active sheets
master = op.load_workbook(
    'Templates/Template_MasterRevenueTracker.xlsx')
master_ws = master.active
proj_tracker = op.load_workbook(
    'Templates/Template_ProjectTracker.xlsx')
track_ws = proj_tracker.active
entry = op.load_workbook('Templates/Template_DailyInput_Example.xlsx')
entry_ws = entry.active

# Create Project Variables
months = ['NONE', 'JAN', 'FEB', 'MAR', 'APR', 'MAY',
          'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
hhs_num = entry_ws['B1'].value  # HHS Project Number
cus_job = entry_ws['B2'].value  # Customer Project Number
month = entry_ws['D5'].value  # Month on input sheet
month_name = months[int(month)]  # Month by name
day = entry_ws['D6'].value  # Day on input sheet
ncn = (track_ws.max_column + 1)  # Next open column number
ncl = op.utils.get_column_letter(ncn)  # Next column letter
proj_info = []  # Project information list
proj_data = []  # Empty data list
day_total = entry_ws['D98'].value  # Daily Total (Dollars)
# Gather input wb data into lists
for row in entry_ws.rows:
    proj_info.append(row[1].value)
proj_info = proj_info[:4]
for row in entry_ws.rows:
    proj_data.append(row[3].value)
proj_data = proj_data[4:98]
# Enter data from lists to project tracker
for i, item in enumerate(proj_data):
    track_ws[f'{ncl}{i+5}'] = item
track_ws[f'{ncl}98'] = day_total
proj_total = track_ws['B99'].value
# Variables for adding proj to Master Tracker
x = 6
filled = True
proj_list = []
# List existing projects and verify does not exist
for row in master_ws.rows:
    proj_list.append(row[1].value)
if str(hhs_num) in proj_list:
    filled = False
# TODO Add months total to project if exists
proj_cell = master_ws['B8']
proj_rev_cell = proj_cell.offset(month, 0)
# Add project and total if does not exist
while filled:
    print(master_ws[f'B{x}'].value)
    if master_ws[f'B{x}'].value != '##-#####':
        x += 1
    else:
        master_ws[f'B{x}'] = hhs_num
        master_ws[f'C{x}'] = proj_total
        filled = False

# TESTING AREA
print(f'New Col Number: {ncn} Letter: {ncl}')
print(f'HHS Proj Number: {hhs_num}')
print(f'Infin Number: {proj_info[1]}')
print(f'Location: {proj_info[2]}')
print(f'Hub Number: {proj_info[-1]}')
print(f'Date on form: {month}-{day}')
print(f'Month on form: {month_name}')
print(f'First data entry: {proj_data[0]}')
print(f'Last data entry: {proj_data[-1]}')
print(f'Length of data: {len(proj_data)}')
print(f'Day Total: {day_total}')
print(f'Project Cell: {proj_cell} Proj Total Cell: {proj_rev_cell}')

# Save Project Tracker
proj_tracker.save(f'Templates/Template_ProjectTracker.xlsx')
entry.save(f'Templates/{hhs_num}_{month}-{day}.xlsx')
master.save(f'Templates/Template_MasterRevenueTracker.xlsx')
