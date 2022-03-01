# !Python3                                              A3R0NA$
import openpyxl as op

# Open Project Tracker and Daily Input, load active sheets
prt = op.load_workbook('Test_Reports/Test-ProjectTracker.xlsx', data_only=True)
prt_ws = prt.active
di = op.load_workbook('Test_Inputs/Test-DailyInput.xlsx', data_only=True)
di_ws = di.active

# Create Project Variables
ncn = (prt_ws.max_column + 1)  # Next open column number
ncl = op.utils.get_column_letter(ncn)
hhs_num = di_ws['B1'].value  # HHS Project Number
date = di_ws['D6'].value  # Date on input sheet
proj_info = []  # Empty project info
prod_data = []  # Empty data list
day_total = di_ws['D98'].value  # Daily Total (Dollars)

# Add project information data set
for row in di_ws.rows:
    proj_info.append(row[1].value)
proj_info = proj_info[:3]
# Add new production numbers to data set, removing top empty cells
for row in di_ws.rows:
    prod_data.append(row[3].value)
prod_data = prod_data[5:-20]

# Verify project info in production tracker

# Add production data to new column in Project Tracker
for i, item in enumerate(prod_data):
    prt_ws[f'{ncl}{i+6}'] = item

# TESTING AREA
print(f'New Col Number: {ncn} Letter: {ncl}')
print(f'HHS Proj Number: {hhs_num}')
print(f'Infin Number: {proj_info[1]}')
print(f'Location: {proj_info[2]}')
print(f'Hub Number: {proj_info[-1]}')
print(f'Date on form: {date}')
print(f'First data entry: {prod_data[0]}')
print(f'Length of data: {len(prod_data)}')
print(f'Day Total: {day_total}')

# Save Project Tracker
prt.save('Test_Reports/Test-ProjectTracker.xlsx')
di.save(f'Test_Reports/Test_History/{hhs_num}-{date}.xlsx')
