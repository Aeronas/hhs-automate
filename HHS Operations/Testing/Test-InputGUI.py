# !Python3                                              A3R0NA$
from tkinter import *
import openpyxl as op

'''
{hhs_num}_{cus_job}_{month}-{day}.xlsx
{hhs_num}_{cus_job}_Production-Tracker.xlsx
'''


# Main window
root = Tk()
root.title("Daily Production Input")
root.iconbitmap('Templates/Images/chart-5_111421.ico')
root.geometry('275x175')


# Command functions
def NewInputWB():
    '''Create new entry workbook'''
    global hhs_num
    global cus_job
    global day
    global month
    # TODO Verify todays entry does not exist
    # Load entry template
    new_wb = op.load_workbook('Templates/Template_Input-Ziply_DailyProduction.xlsx')
    # TODO Enter inputs into workbook
    # Save new copy in reports
    new_wb.Save('Reports/History/{hhs_num}_{cus_job}_{month}-{day}.xlsx')
    # TODO Open the new input copy

def AddDayColumn():
    '''Enter the new days data'''
    global hhs_num
    global cus_job
    global day
    global month
    # Load finished input file
    entry_wb = op.load_workbook(f'Reports/History/{hhs_num}_{cus_job}_{month}-{day}.xlsx')
    # Load existing project tracker
    proj_tracker = op.load_workbook('Reports/{hhs_num}_{cus_job}_Production-Tracker.xlsx')
    # TODO Verify date column does not already exist
    # TODO Enter input wb data into new column
    # TODO Save project tracker changes, Close input wb

def CreateNewProject():
    '''Create new project tracker'''
    # TODO Verify project does not already exist
    new_tracker = op.load_workbook('Templates/Template_ProjectTracker.xlsx')
    # TODO Enter project data into project file
    # TODO Save new project tracker in reports


# Input names and boxes
hhs_num_label = Label(root, text='HHS Project Number:')
hhs_num_label.grid(row=0, column=0, columnspan=2)
hhs_num_entry = Entry(root, width=23)
hhs_num_entry.grid(row=0, column=2, columnspan=2)
hhs_num = hhs_num_entry.get()  # TODO Verify proper format info

cus_job_label = Label(root, text='Customer Job Number:')
cus_job_label.grid(row=1, column=0, columnspan=2)
cus_job_entry = Entry(root, width=23)
cus_job_entry.grid(row=1, column=2, columnspan=2)
cus_job = cus_job_entry.get()  # TODO Verify proper format info

loc_label = Label(root, text='Location:')  # TODO Make a dropdown menu?
loc_label.grid(row=2, column=0, columnspan=2)
loc_entry = Entry(root, width=23)
loc_entry.grid(row=2, column=2, columnspan=2)
loc = loc_entry.get()  # TODO Verify proper format info

date_label = Label(root, text='Date: [Day] [Month]')
date_label.grid(row=3, column=0, columnspan=2)
day_entry = Entry(root, width=11)  # TODO Make a dropdown menu?
day_entry.grid(row=3, column=2, pady=6)
month_entry = Entry(root, width=11)  # TODO Make a dropdown menu?
month_entry.grid(row=3, column=3, pady=6)
day = day_entry.get()  # TODO Verify proper format info
month = month_entry.get()  # TODO Verify proper format info


# Action Buttons
make_wb_button = Button(root, text='Create New Report', command=NewInputWB)
make_wb_button.grid(row=4, column=0, columnspan=2)

enter_data_button = Button(root, text='Enter Daily Report', command=AddDayColumn)
enter_data_button.grid(row=4, column=2, columnspan=2)

new_proj_button = Button(root, text='Create Project', command=CreateNewProject)
new_proj_button.grid(row=5, column=1, columnspan=2, pady=20)



# Main logic loop
root.mainloop()
