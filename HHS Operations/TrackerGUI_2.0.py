# !Python3                                              A3R0NA$
import os
from tkinter import *
from tkinter import messagebox
import openpyxl as op
'''
GUI for input and tracking of project production data.
'''

# Create main window
root = Tk()
root.title('Tracker 2.0')
root.geometry('275x175')
root.bitmapimage('Templates/Images/chart-5_111421.ico')


# Core functions
months = ['NONE', 'JAN', 'FEB', 'MAR', 'APR', 'MAY',
          'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']


def VerifyProjInfo(num, m, d):
    '''
    Verifies project number, day and month are entered correctly.
    '''
    if len(num) == 8 and str(num[2]) == '-':
        if int(m) > 0 and int(m) < 13:
            if int(d) > 0 and int(d) < 32:
                return 1
            else:
                return 0
        else:
            return 0
    else:
        return 0

# TODO Create master project tracker and update each daily input


def NewInputWB():
    '''Creates new entry workbook'''
    hhs_num = hhs_num_box.get()
    cus_job = cus_job_box.get()
    loc = loc_box.get()
    day = day_box.get()
    month = month_box.get()
    # month_name = months[int(month)]  # TODO For future sheets
    # Verify inputs are in correct format
    if VerifyProjInfo(hhs_num, month, day):
        # New excel input file name
        new_entry_name = f'{hhs_num}_{cus_job}_{month}-{day}'
        existing_reports = os.listdir('./Reports/History/')
        # Verify todays entry does not exist
        if new_entry_name in existing_reports:
            messagebox.showerror(
                'Error', 'This date input file already exists!')
            pass
        else:
            # Load entry template and add project info
            new_wb = op.load_workbook(
                'Templates/Template_DailyInput_2.0.xlsx')
            new_ws = new_wb.active
            new_ws['B1'] = hhs_num
            new_ws['B2'] = cus_job
            new_ws['B3'] = loc
            # Save new copy in reports
            new_wb.save(f'Reports/History/{new_entry_name}.xlsx')
            # Open the new input copy
            open_now = messagebox.askyesno(
                'Complete', 'Daily entry file created, open it now?')
            if open_now:
                os.system(
                    f'start EXCEL.EXE Reports/History/{new_entry_name}.xlsx')
            else:
                pass
    else:
        messagebox.showerror('Error', 'Incorrect project info formats!')
        pass


# TODO Need month total variable from tracker wb to update master
def EnterDaysInput():
    '''Enters the new days data'''
    hhs_num = hhs_num_box.get()
    cus_job = cus_job_box.get()
    day = day_box.get()
    month = month_box.get()
    # month_name = months[int(month)]  # TODO For future sheets
    # Load projects input file (completed)
    entry_wb = op.load_workbook(
        f'Reports/History/{hhs_num}_{cus_job}_{month}-{day}.xlsx')
    # TODO Change active worksheet to proper month sheet
    entry_ws = entry_wb.active
    # Load existing project tracker
    proj_tracker = op.load_workbook(
        f'Reports/{hhs_num}_{cus_job}_Production-Tracker.xlsx')
    track_ws = proj_tracker.active
    # TODO Verify date column does not already exist (ask to copy over)
    # Create data variables
    ncn = (track_ws.max_column + 1)  # Next open column number
    ncl = op.utils.get_column_letter(ncn)  # Next column letter
    prod_data = []  # Empty data list
    day_total = entry_ws['D98'].value  # Daily Total (Dollars)
    # Gather input wb data into lists
    for row in entry_ws.rows:
        prod_data.append(row[3].value)
    prod_data = prod_data[4:-20]
    # Enter data from lists to project tracker
    for i, item in enumerate(prod_data):
        track_ws[f'{ncl}{i+5}'] = item
    track_ws['D98'] = day_total
    # Save project tracker changes, Close input wb
    proj_tracker.save(f'Reports/{hhs_num}_{cus_job}_Production-Tracker.xlsx')
    # Notify completion
    messagebox.showinfo(
        'Complete', f'{hhs_num} input for {month}/{day} has been added!')


# TODO Create project entry in master project tracker
# TODO Add master tracking sheet and seperate month sheets
def CreateNewProject():
    '''Creates new project tracker'''
    hhs_num = hhs_num_box.get()
    cus_job = cus_job_box.get()
    loc = loc_box.get()
    day = day_box.get()
    month = month_box.get()
    # Verify project format and does not already exist
    if VerifyProjInfo(hhs_num, month, day):
        new_proj_name = f'{hhs_num}_{cus_job}_Production-Tracker'
        existing_projects = os.listdir('./Reports/')
        if new_proj_name not in existing_projects:
            proj_template = op.load_workbook(
                'Templates/Template_ProjectTracker_2.0.xlsx')
            temp_ws = proj_template.active
            # Enter project information into new project wb
            temp_ws['B1'] = hhs_num
            temp_ws['B2'] = cus_job
            temp_ws['B3'] = loc
            # TODO Add question to enter a hub number for Ziply
            # Save new project tracker and notify completetion
            proj_template.save(f'Reports/{new_proj_name}.xlsx')
            messagebox.showinfo(
                'Complete',
                f'Project Tracker {new_proj_name} has been created.')
        else:
            # Project exists error message
            messagebox.showerror('Error', 'Project already exists!')
            pass
    else:
        # Wrong format error message
        messagebox.showerror('Error', 'Verify project info formats!')


# Input names and boxes
hhs_num_label = Label(root, text='HHS Project Number:')
hhs_num_label.grid(row=0, column=0, columnspan=2)
hhs_num_box = StringVar()
hhs_num_entry = Entry(root, width=23, textvariable=hhs_num_box)
hhs_num_entry.grid(row=0, column=2, columnspan=2)

cus_job_label = Label(root, text='Customer Job Number:')
cus_job_label.grid(row=1, column=0, columnspan=2)
cus_job_box = StringVar()
cus_job_entry = Entry(root, width=23, textvariable=cus_job_box)
cus_job_entry.grid(row=1, column=2, columnspan=2)

loc_label = Label(root, text='Location:')
loc_label.grid(row=2, column=0, columnspan=2)
loc_box = StringVar()
loc_entry = Entry(root, width=23, textvariable=loc_box)
loc_entry.grid(row=2, column=2, columnspan=2)

date_label = Label(root, text='Date: [Month] [Day]')
date_label.grid(row=3, column=0, columnspan=2)
month_box = IntVar()
month_entry = Entry(root, width=11, textvariable=month_box)
month_entry.grid(row=3, column=2, pady=6)
day_box = IntVar()
day_entry = Entry(root, width=11, textvariable=day_box)
day_entry.grid(row=3, column=3, pady=6)

# Get all project information from GUI
# hhs_num = hhs_num_box.get()  # TODO Verify proper format info
# cus_job = cus_job_box.get()  # TODO Verify proper format info
# loc = loc_box.get()  # TODO Verify proper format info
# day = day_box.get()  # TODO Verify proper format info
# month = month_box.get()  # TODO Verify proper format info

# Action Buttons
make_wb_button = Button(root, text='Create New Report', command=NewInputWB)
make_wb_button.grid(row=4, column=0, columnspan=2)

enter_data_button = Button(
    root, text='Enter Daily Report', command=EnterDaysInput)
enter_data_button.grid(row=4, column=2, columnspan=2)

new_proj_button = Button(root, text='Create Project', command=CreateNewProject)
new_proj_button.grid(row=5, column=1, columnspan=2, pady=20)


# Run main logic (Open root window)
root.mainloop()
